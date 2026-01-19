import os
import time
import re
import pandas as pd
from python_calamine import CalamineWorkbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D


def normalize_car_code(val, smart=False):
    val = str(val).strip().upper()
    
    if smart:
        # Regex to find patterns like A-123, B-1050, A123, B1050
        # Looks for a letter (A-Z) followed optionally by a dash/space, then digits
        match = re.search(r'([A-Z])[\s_-]*(\d+)', val)
        if match:
            return f"{match.group(1)}{match.group(2)}"
    
    # Fallback / Default: simple cleanup
    return val.replace(' ', '').replace('_', '').replace('.', '').replace('-', '')

def get_desktop_path():
    user_home = os.path.expanduser("~")
    # Check OneDrive paths first
    od_desktop = os.path.join(user_home, "OneDrive", "Desktop")
    if os.path.exists(od_desktop): return od_desktop
    
    od_desktop_tr = os.path.join(user_home, "OneDrive", "Masaüstü")
    if os.path.exists(od_desktop_tr): return od_desktop_tr
    
    # Fallback to standard Desktop
    return os.path.join(user_home, "Desktop")

def read_excel_smart(file_path, log):
    try:
        if not os.path.exists(file_path):
            log(f"Hata: Dosya bulunamadı: {file_path}")
            return pd.DataFrame()

        # Try Calamine first (Fast)
        rows = None
        try:
            wb = CalamineWorkbook.from_path(file_path)
            sheet_name = wb.sheet_names[0]
            rows = wb.get_sheet_by_name(sheet_name).to_python()
        except Exception as cal_err:
            log(f"Hızlı okuma atlandı, klasik motor deneniyor...")
            # Fallback to pandas
            return pd.read_excel(file_path)

        if not rows or len(rows) == 0: 
            return pd.DataFrame()
        
        best_row_idx = 0
        max_score = -1
        search_keywords = ['KAPINO', 'KAPI NO', 'PLAKA', 'DURUM', 'AĞ DURUMU', 'CIHAZ', 'KOD', 'SÜRE', 'SURE', 'ADRES', 'FAKTÖR', 'NUMARASI']
        
        for i in range(min(30, len(rows))):
            row_str = [str(x).strip().upper() for x in rows[i] if x is not None]
            score = sum(1 for kw in search_keywords if any(kw in val for val in row_str))
            if score > max_score:
                max_score = score
                best_row_idx = i
            if score >= 3: break
        
        headers = [str(x).strip() if x is not None else f"Column_{j}" for j, x in enumerate(rows[best_row_idx])]
        # If headers and data are the same row or something went wrong
        if best_row_idx + 1 >= len(rows):
             return pd.DataFrame(columns=headers)

        df = pd.DataFrame(rows[best_row_idx + 1:], columns=headers)
        return df
    except Exception as e:
        log(f"Dosya okuma kritik hatası: {repr(e)}")
        try:
            return pd.read_excel(file_path)
        except:
            return pd.DataFrame()

def find_best_column(columns, target_keywords):
    for col in columns:
        norm_col = str(col).upper().replace('İ', 'I').replace('Ğ', 'G').replace('Ü', 'U').replace('Ş', 'S').replace('Ö', 'O').replace('Ç', 'C').replace(' ', '').replace('_', '')
        for kw in target_keywords:
            if kw.upper().replace(' ', '') in norm_col: return col
    return None

def generate_premium_excel(df, title, output_filename, log):
    output_path = os.path.join(get_desktop_path(), "IETT Veri Panel Çıktı", output_filename)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Rapor"
    
    # Header Layout: A (Logo), B-C (Title), D (Logo or Last Column)
    num_cols = len(df.columns)
    max_header_col = max(num_cols, 4)
    last_col_letter = chr(64 + max_header_col)
    
    # Column Widths
    for i in range(1, max_header_col + 1):
        ws.column_dimensions[chr(64+i)].width = 25
    
    ws.row_dimensions[1].height = 78
    row_h_px = 104

    def add_logo(path, col_letter, col_idx):
        if os.path.exists(path):
            try:
                img = Image(path)
                target_h = 85
                ratio = target_h / img.height
                img_w = int(img.width * ratio)
                img.width = img_w
                img.height = target_h
                
                col_w_px = ws.column_dimensions[col_letter].width * 7.5
                off_x = pixels_to_EMU(int(max(0, (col_w_px - img_w) / 2)))
                off_y = pixels_to_EMU(int(max(0, (row_h_px - target_h) / 2)))
                marker = AnchorMarker(col=col_idx, colOff=off_x, row=0, rowOff=off_y)
                img.anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(pixels_to_EMU(img_w), pixels_to_EMU(target_h)))
                ws.add_image(img)
            except: pass

    assets_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'public')
    add_logo(os.path.join(assets_dir, 'ibb_logo.png'), 'A', 0)
    add_logo(os.path.join(assets_dir, 'iett_logo.png'), last_col_letter, max_header_col - 1)

    # Title - Merge B1 until LastColumn-1
    if max_header_col > 2:
        title_range = f"B1:{chr(64 + max_header_col - 1)}1"
        ws.merge_cells(title_range)
        title_cell = ws['B1']
    else:
        title_cell = ws['A1']

    title_cell.value = title.upper()
    red_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    title_cell.fill = red_fill
    title_cell.font = Font(color="FFFFFF", bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Apply border and fill to all merged cells
    border_thin = Side(style='thin')
    full_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    if max_header_col > 2:
        for row in ws[title_range]:
            for cell in row:
                cell.border = full_border
                cell.fill = red_fill

    # Date
    date_cell = ws[f"{last_col_letter}2"]
    date_cell.value = f"Rapor Tarihi: {time.strftime('%d.%m.%Y %H:%M')}"
    date_cell.alignment = Alignment(horizontal='right')
    date_cell.font = Font(size=9, italic=True)

    # Table Header (Row 4)
    header_fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for i, h in enumerate(df.columns, 1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = full_border

    # Data (Row 5+)
    for r_idx, (idx, row_data) in enumerate(df.iterrows(), 5):
        # If the row is completely empty, skip it to create a visual gap in Excel
        if all(str(v).strip() == "" or v is None for v in row_data):
            continue
            
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = full_border
            
            # Special Formatting: Red background for '0' in door counts
            col_name = str(df.columns[c_idx-1])
            if "Kapı (Adet)" in col_name and str(val) == "0":
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)

    wb.save(output_path)
    return output_path
